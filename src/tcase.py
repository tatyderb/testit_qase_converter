"""
Описание одного теста.
"""
import json
import sys
from dataclasses import dataclass, field
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import ClassVar


class TRow:
    """ Row in xlsx worksheet in TestIt import."""
    id: int = 0
    direction: str = ''
    suite: str = ''
    case: str = ''
    automated: bool = False
    preconditions: str = ''
    steps: str = ''
    postconditions: str = ''
    result: str = ''
    testdata: str = ''
    comments: str = ''
    iterations: str = ''
    priority: str = 'Medium'
    state: str = 'NotReady'
    tags: list = field(default_factory=list)
    is_empty: bool = False    # True если все ячейки пустые, бывает когда добавляют пустой шаг

    @staticmethod
    def parse_headers_from_str():
        testit_xlsx_headers = 'Id	Direction	Section	TestCaseName	Automated	Preconditions	\
            Steps	Postconditions	ExpectedResult	TestData	Comments	Iterations	Priority	State	\
            CreatedDate	CreatedById	Tags'
        res = {name: index for index, name in enumerate(testit_xlsx_headers.split(), 1)}
        print(res)
        return res

    testit_xlsx_headers = parse_headers_from_str()
    # {'Id': 1, 'Direction': 2, ...}

    @staticmethod
    def get_bool(value: str):
        return True if value == 'true' or value == 'True' else False

    @staticmethod
    def get_words(value: str | None):
        if value is None:
            return []
        return value.split()

    @classmethod
    def get_row(cls, ws: Worksheet, row_index):
        row = cls()
        row.id = ws.cell(row=row_index, column=cls.testit_xlsx_headers['Id']).value
        row.direction = ws.cell(row=row_index, column=cls.testit_xlsx_headers['Direction']).value
        row.suite = ws.cell(row=row_index, column=cls.testit_xlsx_headers['Section']).value
        row.case = ws.cell(row=row_index, column=cls.testit_xlsx_headers['TestCaseName']).value
        row.automated = cls.get_bool(ws.cell(row=row_index, column=cls.testit_xlsx_headers['Automated']).value)
        row.preconditions = ws.cell(row=row_index, column=cls.testit_xlsx_headers['Preconditions']).value
        row.steps = ws.cell(row=row_index, column=cls.testit_xlsx_headers['Steps']).value
        row.postconditions = ws.cell(row=row_index, column=cls.testit_xlsx_headers['Postconditions']).value
        row.result = ws.cell(row=row_index, column=cls.testit_xlsx_headers['ExpectedResult']).value
        row.testdata = ws.cell(row=row_index, column=cls.testit_xlsx_headers['TestData']).value
        row.comment = ws.cell(row=row_index, column=cls.testit_xlsx_headers['Comments']).value
        row.iterations = ws.cell(row=row_index, column=cls.testit_xlsx_headers['Iterations']).value
        row.priority = ws.cell(row=row_index, column=cls.testit_xlsx_headers['Priority']).value
        row.tags = cls.get_words(ws.cell(row=row_index, column=cls.testit_xlsx_headers['Tags']).value)

        # полностью пустая строка, такие мы пропускаем
        row.is_empty = True
        for i in range(1, ws.max_column+1):
            if ws.cell(row=row_index, column=i).value:
                row.is_empty = False
                break

        return row


@dataclass
class Action:
    action: str     # do
    result: str     # expected result

    def json(self, position_index: int = 1, null_if_empty: bool = False):
        if null_if_empty and not self.action:
            return None
        d = {
            "position": position_index,
            "action": self.action,
            "expected_result": self.result,
            "data": "",
            "steps": []
        }
        return d


@dataclass
class ActionList:
    actions: list[Action] = field(default_factory=list)

    def json(self):
        return [action.json(position) for position, action in enumerate(self.actions, 1)]


@dataclass
class TCase:
    # эта часть из xlsx TestIt
    id: int
    name: str
    automated: bool = False
    preconditions: str | None = None
    steps: ActionList = field(default_factory=ActionList)
    postconditions: str | None = None
    testdata: str = ''
    comments: str = ''
    iterations: str = ''
    priority: str = 'normal'
    severity: str = 'medium'
    state: str = 'draft'
    tags: list | None = None
    __id: ClassVar[int] = 20       # переменная класса, нужна, чтобы была сквозная нумерация тесткейсов

    @classmethod
    def get_next_id(cls):
        cls.__id += 1
        return cls.__id

    def json(self):
        d = {
            "id": self.id,
            "title": self.name,
            "description": self.name,
            "preconditions": self.preconditions,
            "postconditions": self.postconditions,
            "priority": self.priority,
            "severity": self.severity,
            "type": "functional",
            "behavior": "undefined",
            "automation": "is-not-automated",
            "status": self.state,
            "is_flaky": "no",
            "layer": "unknown",
            "milestone": None,
            "custom_fields": [],
            "steps_type": "classic",
            "steps": [] if self.steps is None else self.steps.json(),
            "tags": self.tags,
            "params": []
        }
        print(f'Step {self.id} ----------------------------------------------------')
        return d

    def add_row(self, row: TRow):
        """ Это вторая и далее строки теста, то есть precondition | step | postcondition + result.
        В qase предусловие и постусловия БЕЗ ожидаемых результатов, так что добавляю результаты в пред/пост условия.
        """
        print(f'{row.preconditions=} {row.steps=} {row.postconditions=} {row.result=}')
        if row.preconditions:
            self.preconditions = row.preconditions
            if row.result:
                self.preconditions += '\n' + row.result
        elif row.postconditions:
            self.postconditions = row.postconditions
            if row.result:
                self.postconditions += '\n' + row.result
        elif row.steps:
            do = row.steps
            if self.steps is None:
                self.steps = ActionList()
            self.steps.actions.append(Action(do, row.result))
        else:
            if row.result:
                do = 'Тут скорее всего была картинка'
                if self.steps is None:
                    self.steps = ActionList()
                self.steps.actions.append(Action(do, row.result))
            else:
                raise ValueError("")

    @staticmethod
    def convert_priority(testit_value: str):
        """
        Convert testit priority value to qase priority
        testit_value: Highest, High, Medium, Low, Lowest
        qase: high, medium, low, undefined
        """
        match testit_value:
            case 'Medium':
                return 'medium'
            case 'Highest' | 'High':
                return 'high'
            case 'Low' | 'Lowest':
                return 'low'
            case _:
                return 'undefined'

    @staticmethod
    def convert_severity(testit_value: str):
        """
        Convert testit priority value to qase severity
        testit_value: Highest, High, Medium, Low, Lowest
        qase: blocker, critical, major, normal, minor, trivial, undefined
        """
        match testit_value:
            case 'Medium':
                return 'normal'
            case 'High':
                return 'major'
            case 'Highest':
                return 'critical'
            case 'Low':
                return 'minor'
            case 'Lowest':
                return 'trivial'
            case _:
                return 'undefined'

    @staticmethod
    def convert_state(testit_value: str):
        """
        Convert testit status value to qase state
        testit_value: Ready, NeedsWork, NotReady
        qase: actual, draft, deprecated
        """
        match testit_value:
            case 'Ready':
                return 'actual'
            case 'NeedsWork' | 'NotReady':
                return 'draft'
            case _:
                return 'draft'

    @classmethod
    def create(cls, row: TRow):
        return TCase(
            id=TCase.get_next_id(),
            name=row.case,
            automated=row.automated,
            testdata=row.testdata,
            comments=row.comments,
            iterations=row.iterations,
            priority=TCase.convert_priority(row.priority),
            severity=TCase.convert_severity(row.priority),
            state=TCase.convert_state(row.state),
            tags=row.tags
            )


@dataclass
class TSuite:
    id: int = 0
    title: str = ''             # xlsx TestCaseName
    description: str = ''       # дублируем с title
    preconditions: str = ''     # предусловия раздела
    suites: list = field(default_factory=list)         # подразделы, в Testit описаны через / в Direction (путь) и Section (конец пути)
    cases: list[TCase] = field(default_factory=list)   # список тест кейсов, по умолчанию пустой
    __id: ClassVar[int] = 10       # переменная класса, нужна, чтобы была сквозная нумерация разделов

    @classmethod
    def get_next_id(cls):
        cls.__id += 1
        return cls.__id

    def json(self):
        d = {
            "id": self.id,  # сквозная нумерация по всем разделам
            "title": self.title,
            "description": self.title,
            "preconditions": self.preconditions,
            "suites": [suite.json() for suite in self.suites],
            "cases": [case.json() for case in self.cases]
        }
        print(f'Suite: {self.title} -----------------------------')
        print(d)
        return d

    def add_case(self, case: TCase):
        """ Добавляем тест в список тестов."""
        self.cases.append(case)

    @classmethod
    def create(cls, row: TRow | None, suite_name: str | None = None):
        """ Создаем раздел по записи в таблице"""
        suite = cls()
        if row:
            suite.title = row.suite if row else 'Root'
            suite.description = row.suite if row else 'Root'
        else:
            suite.title = suite_name
            suite.description = suite_name
        suite.id = cls.get_next_id()

        return suite


class TProject:

    def __init__(self, project_name: str):
        self.suite_tree: TSuite = TSuite.create(None)           # дерево разделов (в Testit без подразделов)
        self.path: dict = {project_name: self.suite_tree}    # быстрый поиск пути к вершине по кортежу названий разделов в пути
        self.case: TCase | None = None                          # текущий case, который мы будем дальше дочитывать, уже в дереве

    def json(self):
        json_project = {
            'suites': [suite.json() for suite in self.suite_tree.suites]
        }
        print(f'PROJECT ====================================================')
        return json_project

    def add_row(self, row: TRow):
        if row.is_empty:
            return
        if row.id:
            # создаем новый кейс
            self.case = TCase.create(row)
            # выясняем в какой раздел добавлять, если нет - создаем
            suite = self.get_suite(row)
            suite.add_case(self.case)
        else:
            # или добавляем информацию к текущему разбираемому кейсу
            self.case.add_row(row)

    def get_suite(self, row: TRow) -> TSuite:
        """ Возвращает ссылку на раздел в дереве разделов.
            Если надо, создает все TSuite по пути.
        """
        directions = row.direction.split('/')
        directions.append(row.suite)
        # если этот раздел уже заведен, вернем его
        suite_path = '/'.join(directions)
        if suite_path in self.path:
            return self.path[suite_path]
        parent = self.suite_tree
        path = directions[0]
        for suite in directions[1:]:
            path = path + '/' + suite
            # идем по узлам к последнему существующему в пути разделу
            if path in self.path:
                parent = self.path[path]
            else:
                # если раздела нет, то создаем его, добавляем в дерево
                parent.suites.append(TSuite.create(row=None, suite_name=suite))
                # теперь этот узел - последний в пути
                parent = parent.suites[-1]
                # и регистрируем в словаре
                self.path[path] = parent
        return parent


def load_xlsx(filename: str):
    """ По xlsx файлу строит json файл."""
    wb = load_workbook(filename=filename)
    ws = wb.active
    print(type(ws), ws.title)
    print(ws.max_column, ws.max_row)
    project: TProject = TProject(ws.title.replace('Project_', ''))
    for row_index in range(2, ws.max_row + 1):
        row: TRow = TRow.get_row(ws, row_index=row_index)
        print(row)
        print(row.__dict__)
        if row.is_empty:
            print(f"EMPTY Row {row_index} skipped .. OK", file=sys.stderr)
        else:
            project.add_row(row)
            print(f"Row {row_index} parsed .. OK", file=sys.stderr)
    return project


if __name__ == '__main__':
    filename = '../xlsx/Test IT - casino web.xlsx'
    filename = '../xlsx/Test IT - 5 draft tests and checklists.xlsx'
    filename = '../xlsx/Test IT - 5 draft tests without checklists.xlsx'
    filename = '../xlsx/Test IT - glory.xlsx'
    filename = '../xlsx/Test IT - 3 draft tests.xlsx'
    project = load_xlsx(filename=filename)
    with open('../json/qase.json', 'w') as fout:
        json.dump(project.json(), fp=fout)
